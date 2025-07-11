"""
Working Admin Routes Implementation
"""

from flask import render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
from models import db, User, Setting, AuditLog, RoleEnum, RolePermission, WorkflowTemplate, WorkflowLibrary, Department, HelpCategory, HelpArticle, NotificationSetting, FrequencyEnum, Problem, Project, OrganizationSettings, WorkflowConfiguration
from werkzeug.security import generate_password_hash
from admin.forms import UserForm
from datetime import datetime
from utils.date import format_datetime
import os

def init_admin_routes(app):
    """Initialize admin routes on Flask app"""
    print(f"🔧 Route Registration Debug: Initializing admin routes on app {app}")
    
    # Test route to verify registration is working
    @app.route('/admin/route-test')
    def admin_route_test():
        return "<h1>Admin Route Test Working!</h1><p>Route registration is successful.</p>"
    
    def admin_required(f):
        def wrapper(*args, **kwargs):
            user = current_user
            print(f"🔧 Admin Required Debug: User authenticated: {user.is_authenticated}")
            if user.is_authenticated:
                print(f"🔧 Admin Required Debug: User role: {user.role}, Role value: {user.role.value}")
            
            if not user.is_authenticated or user.role.value != 'Admin':
                flash('Admin access required', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        wrapper.__name__ = f.__name__
        return wrapper
    
    def log_action(action, details=None):
        """Log admin actions"""
        try:
            user = current_user
            log = AuditLog(
                user_id=user.id if user else None,
                action=action,
                details=details,
                ip_address=getattr(request, 'remote_addr', None) if request else None
            )
            db.session.add(log)
            db.session.commit()
        except:
            pass
    
    @app.route('/admin/')
    @login_required
    @admin_required
    def admin_dashboard():
        # Calculate comprehensive statistics - ORGANIZATION FILTERED
        from models import User, Department, Problem, Project, Epic, BusinessCase
        
        # Filter all queries by organization_id for proper multi-tenant isolation
        org_id = current_user.organization_id
        users_count = User.query.filter_by(organization_id=org_id).count()
        departments_count = Department.query.filter_by(organization_id=org_id).count()
        problems_count = Problem.query.filter_by(organization_id=org_id).count()
        business_cases_count = BusinessCase.query.filter_by(organization_id=org_id).count()
        projects_count = Project.query.filter_by(organization_id=org_id).count()
        
        # Calculate pending review counts - ORGANIZATION FILTERED
        try:
            pending_epics = Epic.query.filter_by(status='Submitted', organization_id=org_id).count()
            pending_cases = BusinessCase.query.filter_by(status='Submitted', organization_id=org_id).count()
            pending_projects = Project.query.filter_by(status='Submitted', organization_id=org_id).count()
        except:
            pending_epics = 0
            pending_cases = 0
            pending_projects = 0
        
        stats = {
            'users': users_count,
            'departments': departments_count,
            'problems': problems_count,
            'business_cases': business_cases_count,
            'projects': projects_count,
            'pending_epics': pending_epics,
            'pending_cases': pending_cases,
            'pending_projects': pending_projects,
            'total_pending': pending_epics + pending_cases + pending_projects
        }
        
        # Create pending_metrics for template compatibility
        pending_metrics = {
            'pending_epics': pending_epics,
            'pending_cases': pending_cases,
            'pending_projects': pending_projects,
            'total_pending': pending_epics + pending_cases + pending_projects
        }
        
        # Calculate role distribution - ORGANIZATION FILTERED
        try:
            from sqlalchemy import func
            role_counts = db.session.query(User.role, func.count(User.role)).filter_by(organization_id=org_id).group_by(User.role).all()
            role_distribution = {role.value: count for role, count in role_counts}
        except:
            role_distribution = {}
        
        # Create health metrics for template
        health_metrics = {
            'active_notifications': 0,
            'failed_triage_actions': 0,
            'system_health': 'good',
            'error_rate': 0,
            'uptime': '99.9%'
        }
        
        # Create empty alerts and triage_activity for template
        alerts = []
        triage_activity = []
        
        try:
            recent_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(5).all()
        except:
            recent_logs = []
        return render_template('admin/dashboard.html', 
                             stats=stats, 
                             pending_metrics=pending_metrics, 
                             role_distribution=role_distribution,
                             health_metrics=health_metrics,
                             alerts=alerts,
                             triage_activity=triage_activity,
                             recent_activity=recent_logs)
    
    @app.route('/admin/settings', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_settings():
        if request.method == 'POST':
            key = request.form.get('key')
            value = request.form.get('value')
            description = request.form.get('description')
            
            if key and value:
                setting = Setting(key=key, value=value, description=description)
                db.session.add(setting)
                db.session.commit()
                log_action('CREATE_SETTING', f'Created {key}')
                flash('Setting created successfully', 'success')
            
            return redirect(url_for('admin_settings'))
        
        settings = Setting.query.all()
        return render_template('admin/settings.html', settings=settings)
    
    @app.route('/admin/settings/<setting_id>', methods=['POST'])
    @login_required
    @admin_required
    def admin_update_setting(setting_id):
        setting = Setting.query.get_or_404(setting_id)
        setting.value = request.form.get('value', setting.value)
        setting.description = request.form.get('description', setting.description)
        setting.updated_at = datetime.utcnow()
        db.session.commit()
        log_action('UPDATE_SETTING', f'Updated {setting.key}')
        flash('Setting updated successfully', 'success')
        return redirect(url_for('admin_settings'))
    
    @app.route('/admin/settings/<setting_id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def admin_delete_setting(setting_id):
        setting = Setting.query.get_or_404(setting_id)
        key = setting.key
        db.session.delete(setting)
        db.session.commit()
        log_action('DELETE_SETTING', f'Deleted {key}')
        flash('Setting deleted successfully', 'success')
        return redirect(url_for('admin_settings'))
    
    @app.route('/admin/users')
    @login_required
    @admin_required
    def admin_users():
        page = request.args.get('page', 1, type=int)
        # Get search and filter parameters
        search = request.args.get('search', '')
        department_filter = request.args.get('department', '')
        role_filter = request.args.get('role', '')
        
        # Build query with department scoping
        # CRITICAL: Filter by organization to ensure data isolation
        query = User.query.filter(User.organization_id == current_user.organization_id)
        
        # Department scoping for non-Admin users (Directors can only manage their department hierarchy)
        if current_user.role.value != 'Admin':
            if current_user.department_id:
                allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
                query = query.filter(User.department_id.in_(allowed_dept_ids))
        
        if search:
            query = query.filter(db.or_(
                User.name.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%')
            ))
        
        if department_filter:
            query = query.filter(User.department_id == department_filter)
            
        if role_filter:
            query = query.filter(User.role == role_filter)
        
        users = query.all()
        
        # Also scope departments dropdown for non-Admin users
        # Note: Department model has organization_id filtering
        if current_user.role.value != 'Admin' and current_user.department_id:
            allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
            departments = Department.query.filter(Department.id.in_(allowed_dept_ids)).all()
        else:
            # Get all departments for this organization
            departments = Department.query.filter_by(organization_id=current_user.organization_id).all()
        
        return render_template('admin/users.html', 
                             users=users, 
                             departments=departments,
                             search=search,
                             department_filter=department_filter,
                             role_filter=role_filter)
    
    @app.route('/admin/users/create', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_create_user():
        if request.method == 'POST':
            name = request.form.get('name')
            email = request.form.get('email')
            role = request.form.get('role')
            department_id = request.form.get('department_id')
            
            if name and email and role:
                # Check if user already exists
                existing_user = User.query.filter_by(email=email).first()
                if existing_user:
                    flash('User with this email already exists', 'error')
                    return redirect(url_for('admin_create_user'))
                
                from werkzeug.security import generate_password_hash
                user = User(
                    name=name,
                    email=email,
                    role=RoleEnum(role),
                    password_hash=generate_password_hash('temppass123'),
                    is_active=True
                )
                
                # Set department if provided and validate access for non-Admin users
                if department_id:
                    dept_id = int(department_id)
                    
                    # Department access validation for non-Admin users
                    if current_user.role.value != 'Admin':
                        if current_user.department_id:
                            allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
                            if dept_id not in allowed_dept_ids:
                                flash('You can only create users in your department hierarchy', 'error')
                                return redirect(url_for('admin_create_user'))
                    
                    user.department_id = dept_id
                    
                db.session.add(user)
                db.session.commit()
                log_action('CREATE_USER', f'Created user {email}')
                flash('User created successfully', 'success')
                return redirect(url_for('admin_users'))
        
        # Scope departments for non-Admin users
        if current_user.role.value != 'Admin' and current_user.department_id:
            allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
            departments = Department.query.filter(Department.id.in_(allowed_dept_ids)).all()
        else:
            departments = Department.query.filter_by(organization_id=current_user.organization_id).all()
        roles = list(RoleEnum)
        return render_template('admin/user_form.html', 
                             departments=departments, 
                             roles=roles, 
                             user=None)
    
    @app.route('/admin/users/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_edit_user(id):
        user = User.query.get_or_404(id)
        
        # Department access validation for non-Admin users
        if current_user.role.value != 'Admin':
            if current_user.department_id and user.department_id:
                allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
                if user.department_id not in allowed_dept_ids:
                    flash('You can only edit users in your department hierarchy', 'error')
                    return redirect(url_for('admin_users'))
        
        if request.method == 'POST':
            user.name = request.form.get('name', user.name)
            user.email = request.form.get('email', user.email)
            role_value = request.form.get('role')
            if role_value:
                user.role = RoleEnum(role_value)
            department_id = request.form.get('department_id')
            if department_id:
                dept_id = int(department_id)
                
                # Department access validation for non-Admin users
                if current_user.role.value != 'Admin':
                    if current_user.department_id:
                        allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
                        if dept_id not in allowed_dept_ids:
                            flash('You can only assign users to departments in your hierarchy', 'error')
                            return redirect(url_for('admin_edit_user', id=id))
                
                user.department_id = dept_id
            else:
                user.department_id = None
            
            db.session.commit()
            log_action('UPDATE_USER', f'Updated user {user.email}')
            flash('User updated successfully', 'success')
            return redirect(url_for('admin_users'))
        
        # Scope departments for non-Admin users
        if current_user.role.value != 'Admin' and current_user.department_id:
            allowed_dept_ids = current_user.department.get_descendant_ids(include_self=True)
            departments = Department.query.filter(Department.id.in_(allowed_dept_ids)).all()
        else:
            departments = Department.query.filter_by(organization_id=current_user.organization_id).all()
        roles = list(RoleEnum)
        
        # Create proper form with user data
        form = UserForm()
        
        # Set department choices manually since Department.get_hierarchical_choices() might not exist
        dept_choices = [(0, 'No Department')] + [(d.id, d.name) for d in departments]
        form.department.choices = dept_choices
        
        if user:
            form.name.data = user.name
            form.email.data = user.email
            form.role.data = user.role.name if user.role else ''
            form.department.data = user.department_id
            form.is_active.data = user.is_active
        
        return render_template('admin/user_form.html', 
                             form=form,
                             user=user, 
                             departments=departments, 
                             roles=roles)
    
    @app.route('/admin/users/<int:id>/toggle', methods=['POST'])
    @login_required
    @admin_required
    def admin_toggle_user_status(id):
        user = User.query.get_or_404(id)
        user.is_active = not user.is_active
        db.session.commit()
        
        status = 'activated' if user.is_active else 'deactivated'
        log_action('TOGGLE_USER_STATUS', f'User {user.email} {status}')
        flash(f'User {status} successfully', 'success')
        return redirect(url_for('admin_users'))
    
    @app.route('/admin/audit-logs')
    @login_required
    @admin_required
    def admin_audit_logs():
        page = request.args.get('page', 1, type=int)
        try:
            logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).paginate(
                page=page, per_page=50, error_out=False)
        except:
            # Handle case where AuditLog table might not exist
            logs = type('obj', (object,), {'items': [], 'pages': 0, 'has_prev': False, 'has_next': False, 'total': 0})()
        users = User.query.all()
        return render_template('admin/audit_logs.html', logs=logs, users=users)
    
    @app.route('/admin/roles', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_role_permissions():
        """Admin role permissions management"""
        if request.method == 'POST':
            # Handle adding new role permission
            role = request.form.get('role')
            module = request.form.get('module')
            can_create = 'can_create' in request.form
            can_read = 'can_read' in request.form
            can_update = 'can_update' in request.form
            can_delete = 'can_delete' in request.form
            
            # Check if role permission already exists
            existing = RolePermission.query.filter_by(role=RoleEnum(role), module=module).first()
            if existing:
                # Update existing permission
                existing.can_create = can_create
                existing.can_read = can_read
                existing.can_update = can_update
                existing.can_delete = can_delete
                log_action(f"Updated role permission for {role} - {module}")
            else:
                # Create new permission
                new_permission = RolePermission(
                    role=RoleEnum(role),
                    module=module,
                    can_create=can_create,
                    can_read=can_read,
                    can_update=can_update,
                    can_delete=can_delete
                )
                db.session.add(new_permission)
                log_action(f"Created role permission for {role} - {module}")
            
            db.session.commit()
            return redirect(url_for('admin_role_permissions'))
        
        roles = RolePermission.query.all()
        
        # Define available modules based on actual system models
        available_modules = [
            'Problem', 'BusinessCase', 'Project', 'User', 
            'Department', 'Notification', 'Setting', 'Report'
        ]
        
        log_action("Viewed role permissions")
        return render_template('admin/role_permissions.html', roles=roles, available_modules=available_modules)
    
    @app.route('/admin/roles/<int:role_id>/edit', methods=['POST'])
    @login_required
    @admin_required
    def edit_role_permission(role_id):
        """Edit existing role permission"""
        permission = RolePermission.query.get_or_404(role_id)
        
        permission.can_create = 'can_create' in request.form
        permission.can_read = 'can_read' in request.form
        permission.can_update = 'can_update' in request.form
        permission.can_delete = 'can_delete' in request.form
        
        db.session.commit()
        log_action(f"Updated role permission {permission.role.value} - {permission.module}")
        
        return redirect(url_for('admin_role_permissions'))
    
    @app.route('/admin/roles/<int:role_id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def delete_role_permission(role_id):
        """Delete role permission"""
        permission = RolePermission.query.get_or_404(role_id)
        role_name = f"{permission.role.value} - {permission.module}"
        
        db.session.delete(permission)
        db.session.commit()
        log_action(f"Deleted role permission {role_name}")
        
        return redirect(url_for('admin_role_permissions'))
    
    @app.route('/admin/workflows', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_workflows():
        """Admin workflow templates management - redirects to working version"""
        # Always redirect to the working version to avoid transaction issues
        return redirect(url_for('admin_workflows_fixed'))
    
    @app.route('/admin/workflows/import/<int:library_id>', methods=['POST'])
    @login_required
    @admin_required
    def import_workflow_from_library(library_id):
        """Import workflow from library and create customizable template"""
        try:
            # Get the library workflow
            library_workflow = WorkflowLibrary.query.get_or_404(library_id)
            
            # Create a new workflow template based on the library workflow
            imported_name = f"{library_workflow.name} (Imported)"
            
            # Check if this name already exists for this organization
            counter = 1
            base_name = imported_name
            while WorkflowTemplate.query.filter_by(
                name=imported_name, 
                organization_id=current_user.organization_id
            ).first():
                counter += 1
                imported_name = f"{base_name} ({counter})"
            
            new_workflow = WorkflowTemplate(
                name=imported_name,
                description=f"Imported from library: {library_workflow.description}",
                definition=library_workflow.definition,  # Copy the JSON definition
                organization_id=current_user.organization_id,
                created_by=current_user.id,
                is_active=True
            )
            
            db.session.add(new_workflow)
            db.session.commit()
            
            flash(f'Successfully imported workflow: {imported_name}', 'success')
            print(f"🔧 Imported workflow '{imported_name}' from library for org {current_user.organization_id}")
            
        except Exception as e:
            print(f"🔧 Error importing workflow: {str(e)}")
            db.session.rollback()
            flash(f'Error importing workflow: {str(e)}', 'error')
        
        return redirect(url_for('admin_workflows'))

    @app.route('/admin/workflows-fixed', methods=['GET'])
    @login_required
    @admin_required
    def admin_workflows_fixed():
        """Fixed workflow templates management that bypasses transaction issues"""
        try:
            # Force close and recreate database session to clear transaction errors
            db.session.close()
            
            # Ensure essential workflows exist for this organization
            _ensure_essential_workflows()
            
            # Use raw SQL to bypass transaction issues  
            from sqlalchemy import text
            result = db.session.execute(
                text("SELECT id, name, description, is_active FROM workflow_templates WHERE organization_id = :org_id ORDER BY name"),
                {"org_id": current_user.organization_id}
            )
            workflow_data = result.fetchall()
            
            # Convert to objects manually
            workflows = []
            for row in workflow_data:
                # Create a proper object with all attributes
                class WorkflowObj:
                    def __init__(self, id, name, description, is_active):
                        self.id = id
                        self.name = name
                        self.description = description
                        self.is_active = is_active
                
                workflow = WorkflowObj(row[0], row[1], row[2], row[3])
                workflows.append(workflow)
            
            print(f"🔧 Loaded {len(workflows)} workflows for org {current_user.organization_id}")
            
            # Debug: print workflow details
            for wf in workflows:
                print(f"🔧 Workflow: {wf.name} (Active: {wf.is_active})")
            
            return render_template('admin/workflows_fixed.html', 
                                 workflows=workflows)
        except Exception as e:
            print(f"🔧 Error loading workflows: {str(e)}")
            import traceback
            traceback.print_exc()
            # Return empty lists if there's an error
            return render_template('admin/workflows_fixed.html', workflows=[])

    def _ensure_essential_workflows():
        """Ensure all essential DeciFrame workflows exist for the organization"""
        from sqlalchemy import text
        
        essential_workflows = [
            {
                'name': 'Problem-to-Business Case Workflow',
                'description': 'Complete workflow from problem identification to business case creation with triage rules',
                'is_active': True
            },
            {
                'name': 'Business Case Review & Approval',
                'description': 'Multi-stage business case review process with BA assignment and ROI validation',
                'is_active': True
            },
            {
                'name': 'Project Milestone Management',
                'description': 'Project milestone tracking with automated status updates and notifications',
                'is_active': True
            },
            {
                'name': 'Epic & Story Management',
                'description': 'AI-powered epic creation and user story breakdown with requirements generation',
                'is_active': True
            },
            {
                'name': 'Solution Recommendation Process',
                'description': 'AI-powered solution engine with implementation workflow integration',
                'is_active': True
            },
            {
                'name': 'Department Escalation Workflow',
                'description': 'Hierarchical department escalation with role-based routing and approval chains',
                'is_active': True
            },
            {
                'name': 'Notification & Alert Management',
                'description': 'Automated notification system with email integration and workflow triggers',
                'is_active': True
            },
            {
                'name': 'Business Case to Project Conversion',
                'description': 'Automated conversion from approved business cases to project management with milestone creation',
                'is_active': True
            }
        ]
        
        try:
            # Check which workflows already exist
            existing_result = db.session.execute(
                text("SELECT name FROM workflow_templates WHERE organization_id = :org_id"),
                {"org_id": current_user.organization_id}
            )
            existing_names = [row[0] for row in existing_result.fetchall()]
            
            # Create missing workflows
            for workflow in essential_workflows:
                if workflow['name'] not in existing_names:
                    db.session.execute(
                        text("""INSERT INTO workflow_templates (name, description, is_active, organization_id, definition, created_by) 
                                VALUES (:name, :description, :is_active, :org_id, :definition, :created_by)"""),
                        {
                            'name': workflow['name'],
                            'description': workflow['description'],
                            'is_active': workflow['is_active'],
                            'org_id': current_user.organization_id,
                            'definition': '{"type": "standard", "configurable": true}',
                            'created_by': current_user.id
                        }
                    )
                    print(f"🔧 Created essential workflow: {workflow['name']}")
            
            db.session.commit()
            
        except Exception as e:
            print(f"🔧 Error ensuring essential workflows: {str(e)}")
            db.session.rollback()

    @app.route('/admin/workflows-import', methods=['POST'])
    @login_required
    @admin_required
    def import_workflow_template():
        """Simple template import that creates a workflow from predefined templates"""
        template_name = request.form.get('template_name')
        
        # DeciFrame-specific workflow template definitions
        template_definitions = {
            'Problem-to-Business Case Workflow': {
                'description': 'Complete workflow from problem identification to business case creation with triage rules',
                'definition': {"steps": [
                    {"id": "step1", "name": "Problem Submission", "type": "task", "assignee": "Staff", "duration": "30 minutes", "module": "problems"},
                    {"id": "step2", "name": "Auto-Classification", "type": "automated", "assignee": "AI_System", "duration": "immediate", "module": "ai"},
                    {"id": "step3", "name": "Triage Review", "type": "approval", "assignee": "Manager", "duration": "2 days", "module": "problems"},
                    {"id": "step4", "name": "Business Case Creation", "type": "task", "assignee": "BA", "duration": "3 days", "module": "business"},
                    {"id": "step5", "name": "Initial ROI Analysis", "type": "task", "assignee": "BA", "duration": "1 day", "module": "business"}
                ]}
            },
            'Business Case Review & Approval': {
                'description': 'Multi-stage business case review process with BA assignment and ROI validation',
                'definition': {"steps": [
                    {"id": "step1", "name": "BA Assignment", "type": "assignment", "assignee": "Manager", "duration": "1 day", "module": "business"},
                    {"id": "step2", "name": "Light Case Analysis", "type": "task", "assignee": "BA", "duration": "2 days", "module": "business"},
                    {"id": "step3", "name": "Full Case Development", "type": "conditional", "condition": "ROI > $25k", "assignee": "BA", "duration": "5 days", "module": "business"},
                    {"id": "step4", "name": "Director Review", "type": "approval", "assignee": "Director", "duration": "3 days", "module": "business"},
                    {"id": "step5", "name": "CEO Approval", "type": "conditional", "condition": "budget > $100k", "assignee": "CEO", "duration": "2 days", "module": "business"}
                ]}
            },
            'Epic & Story Management': {
                'description': 'AI-powered epic creation and user story breakdown with requirements generation',
                'definition': {"steps": [
                    {"id": "step1", "name": "Epic Creation", "type": "task", "assignee": "BA", "duration": "2 days", "module": "business"},
                    {"id": "step2", "name": "AI Story Generation", "type": "automated", "assignee": "AI_System", "duration": "30 minutes", "module": "ai"},
                    {"id": "step3", "name": "Story Refinement", "type": "task", "assignee": "BA", "duration": "1 day", "module": "business"},
                    {"id": "step4", "name": "Epic Sync Review", "type": "approval", "assignee": "PM", "duration": "1 day", "module": "business"},
                    {"id": "step5", "name": "Requirements Export", "type": "automated", "assignee": "System", "duration": "immediate", "module": "business"}
                ]}
            },
            'Business Case to Project Conversion': {
                'description': 'Automated conversion from approved business cases to project management with milestone creation',
                'definition': {"steps": [
                    {"id": "step1", "name": "Project Creation", "type": "automated", "assignee": "System", "duration": "immediate", "module": "projects"},
                    {"id": "step2", "name": "PM Assignment", "type": "assignment", "assignee": "Director", "duration": "1 day", "module": "projects"},
                    {"id": "step3", "name": "Milestone Planning", "type": "task", "assignee": "PM", "duration": "2 days", "module": "projects"},
                    {"id": "step4", "name": "Resource Allocation", "type": "task", "assignee": "PM", "duration": "1 day", "module": "projects"},
                    {"id": "step5", "name": "Project Kickoff", "type": "notification", "assignee": "PM", "duration": "1 day", "module": "projects"}
                ]}
            },
            'Project Review & Milestone Tracking': {
                'description': 'Project milestone monitoring with automated notifications and success prediction',
                'definition': {"steps": [
                    {"id": "step1", "name": "Milestone Check", "type": "automated", "assignee": "System", "duration": "daily", "module": "projects"},
                    {"id": "step2", "name": "Success Prediction", "type": "automated", "assignee": "ML_System", "duration": "immediate", "module": "predict"},
                    {"id": "step3", "name": "Risk Assessment", "type": "task", "assignee": "PM", "duration": "1 day", "module": "projects"},
                    {"id": "step4", "name": "Stakeholder Updates", "type": "notification", "assignee": "PM", "duration": "immediate", "module": "notifications"},
                    {"id": "step5", "name": "Course Correction", "type": "conditional", "condition": "risk_high", "assignee": "PM", "duration": "2 days", "module": "projects"}
                ]}
            },
            'Solution Recommendation Process': {
                'description': 'AI-powered solution recommendation engine with implementation workflow',
                'definition': {"steps": [
                    {"id": "step1", "name": "Problem Analysis", "type": "task", "assignee": "BA", "duration": "1 day", "module": "solutions"},
                    {"id": "step2", "name": "AI Recommendation", "type": "automated", "assignee": "AI_System", "duration": "immediate", "module": "solutions"},
                    {"id": "step3", "name": "Solution Evaluation", "type": "task", "assignee": "BA", "duration": "2 days", "module": "solutions"},
                    {"id": "step4", "name": "Implementation Planning", "type": "task", "assignee": "PM", "duration": "3 days", "module": "solutions"},
                    {"id": "step5", "name": "Solution Handoff", "type": "approval", "assignee": "Director", "duration": "1 day", "module": "solutions"}
                ]}
            },
            'Department Escalation Workflow': {
                'description': 'Hierarchical department escalation with role-based routing and approval chains',
                'definition': {"steps": [
                    {"id": "step1", "name": "Initial Assignment", "type": "automated", "assignee": "System", "duration": "immediate", "module": "dept"},
                    {"id": "step2", "name": "Department Review", "type": "task", "assignee": "Manager", "duration": "2 days", "module": "dept"},
                    {"id": "step3", "name": "Escalation Check", "type": "conditional", "condition": "unresolved", "assignee": "System", "duration": "immediate", "module": "dept"},
                    {"id": "step4", "name": "Senior Management", "type": "approval", "assignee": "Director", "duration": "1 day", "module": "dept"},
                    {"id": "step5", "name": "Executive Review", "type": "conditional", "condition": "critical", "assignee": "CEO", "duration": "immediate", "module": "dept"}
                ]}
            },
            'Notification & Alert Management': {
                'description': 'Automated notification system with email integration and workflow triggers',
                'definition': {"steps": [
                    {"id": "step1", "name": "Event Detection", "type": "automated", "assignee": "System", "duration": "immediate", "module": "notifications"},
                    {"id": "step2", "name": "Rule Evaluation", "type": "automated", "assignee": "System", "duration": "immediate", "module": "notifications"},
                    {"id": "step3", "name": "Notification Routing", "type": "automated", "assignee": "System", "duration": "immediate", "module": "notifications"},
                    {"id": "step4", "name": "Email Delivery", "type": "automated", "assignee": "SendGrid", "duration": "immediate", "module": "notifications"},
                    {"id": "step5", "name": "Delivery Tracking", "type": "automated", "assignee": "System", "duration": "immediate", "module": "notifications"}
                ]}
            }
        }
        
        if template_name in template_definitions:
            try:
                from models import WorkflowTemplate
                
                template_data = template_definitions[template_name]
                imported_name = f"{template_name} (Imported)"
                
                # Check for naming conflicts
                counter = 1
                base_name = imported_name
                while WorkflowTemplate.query.filter_by(
                    name=imported_name, 
                    organization_id=current_user.organization_id
                ).first():
                    counter += 1
                    imported_name = f"{base_name} ({counter})"
                
                new_workflow = WorkflowTemplate(
                    name=imported_name,
                    description=template_data['description'],
                    definition=template_data['definition'],
                    organization_id=current_user.organization_id,
                    created_by=current_user.id,
                    is_active=True
                )
                
                db.session.add(new_workflow)
                db.session.commit()
                
                flash(f'Successfully imported workflow: {imported_name}', 'success')
                print(f"🔧 Imported workflow '{imported_name}' for org {current_user.organization_id}")
                
            except Exception as e:
                print(f"🔧 Error importing workflow: {str(e)}")
                db.session.rollback()
                flash(f'Error importing workflow: {str(e)}', 'error')
        else:
            flash(f'Template "{template_name}" not found', 'error')
        
        return redirect(url_for('admin_workflows_fixed'))

    @app.route('/admin/workflows/<int:workflow_id>/toggle', methods=['POST'])
    @login_required
    @admin_required
    def toggle_workflow(workflow_id):
        """Toggle workflow template active status"""
        try:
            workflow = WorkflowTemplate.query.filter_by(
                id=workflow_id, 
                organization_id=current_user.organization_id
            ).first_or_404()
            
            workflow.is_active = not workflow.is_active
            db.session.commit()
            
            status = "activated" if workflow.is_active else "deactivated"
            # Skip audit logging to avoid transaction issues
            print(f"🔧 Workflow template {workflow.name} {status}")
            
            return jsonify({'success': True, 'message': f'Workflow {status} successfully'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': f'Failed to toggle workflow: {str(e)}'}), 500
    
    @app.route('/admin/workflows/<int:workflow_id>/configure', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def configure_workflow(workflow_id):
        """Configure workflow parameters (Tier 1 customization)"""
        workflow = WorkflowTemplate.query.filter_by(
            id=workflow_id, 
            organization_id=current_user.organization_id
        ).first_or_404()
        
        if request.method == 'POST':
            try:
                # Update workflow active status
                workflow.is_active = 'workflow_active' in request.form
                
                # Get or create configuration
                config = WorkflowConfiguration.get_or_create_for_workflow(
                    current_user.organization_id,
                    workflow_id,
                    current_user.id
                )
                
                # Update business case configuration
                config.full_case_threshold = float(request.form.get('full_case_threshold', 25000.0))
                config.ba_assignment_timeout = int(request.form.get('ba_assignment_timeout', 48))
                config.director_approval_timeout = int(request.form.get('director_approval_timeout', 72))
                
                # Update notification configuration
                config.reminder_frequency = int(request.form.get('reminder_frequency', 24))
                config.escalation_levels = int(request.form.get('escalation_levels', 3))
                config.email_notifications = 'email_notifications' in request.form
                config.sms_notifications = 'sms_notifications' in request.form
                
                # Update project configuration
                config.milestone_reminder_days = int(request.form.get('milestone_reminder_days', 7))
                config.overdue_escalation_days = int(request.form.get('overdue_escalation_days', 3))
                
                # Update problem configuration
                config.auto_triage_enabled = 'auto_triage_enabled' in request.form
                config.high_priority_escalation_hours = int(request.form.get('high_priority_escalation_hours', 4))
                config.problem_resolution_sla = int(request.form.get('problem_resolution_sla', 72))
                
                # Update workflow steps configuration
                config.skip_ba_assignment = 'skip_ba_assignment' in request.form
                config.require_manager_approval = 'require_manager_approval' in request.form
                config.enable_peer_review = 'enable_peer_review' in request.form
                
                # Update role configuration
                assignee_roles = request.form.getlist('assignee_roles')
                approval_roles = request.form.getlist('approval_roles')
                config.set_assignee_roles(assignee_roles)
                config.set_approval_roles(approval_roles)
                
                db.session.commit()
                flash(f'Configuration updated for workflow: {workflow.name}', 'success')
                
                # Log configuration change
                print(f"🔧 Workflow configuration updated for {workflow.name} by {current_user.email}")
                
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating configuration: {str(e)}', 'error')
                print(f"🔧 Error updating workflow configuration: {str(e)}")
            
            return redirect(url_for('admin_workflows_fixed'))
        
        # GET request - show configuration form
        config = WorkflowConfiguration.get_or_create_for_workflow(
            current_user.organization_id,
            workflow_id,
            current_user.id
        )
        
        return render_template('admin/workflow_configuration.html', 
                             workflow=workflow, 
                             config=config)

    @app.route('/admin/workflows/<int:workflow_id>/edit', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def edit_workflow(workflow_id):
        """Edit existing workflow template"""
        workflow = WorkflowTemplate.query.filter_by(id=workflow_id, organization_id=current_user.organization_id).first_or_404()
        
        if request.method == 'GET':
            # Show edit form - redirect to configure workflow for now (Tier 1 implementation)
            return redirect(url_for('configure_workflow', workflow_id=workflow_id))
        
        # Handle POST request - update workflow
        workflow.name = request.form.get('name')
        workflow.description = request.form.get('description')
        workflow.is_active = 'is_active' in request.form
        
        template_content = request.form.get('template_content', '{}')
        try:
            import json
            workflow.definition = json.loads(template_content)
            db.session.commit()
            log_action(f"Updated workflow template: {workflow.name}")
        except json.JSONDecodeError:
            log_action(f"Failed to update workflow template: Invalid JSON")
        
        return redirect(url_for('admin_workflows'))
    
    @app.route('/admin/workflows/<int:workflow_id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def delete_workflow(workflow_id):
        """Delete workflow template"""
        try:
            workflow = WorkflowTemplate.query.filter_by(
                id=workflow_id, 
                organization_id=current_user.organization_id
            ).first_or_404()
            
            workflow_name = workflow.name
            db.session.delete(workflow)
            db.session.commit()
            
            # Skip audit logging to avoid transaction issues
            print(f"🔧 Deleted workflow template: {workflow_name}")
            
            return jsonify({'success': True, 'message': f'Workflow "{workflow_name}" deleted successfully'})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': f'Failed to delete workflow: {str(e)}'}), 500



    # Bulk Data Import Routes - Add URL name for template reference
    @app.route('/admin/import-data', methods=['GET', 'POST'], endpoint='admin_import_data')
    @admin_required
    def admin_import_data():
        """Bulk Data Import wizard main page"""
        from models import ImportJob
        
        # Handle form submission (upload)
        if request.method == 'POST':
            try:
                from werkzeug.utils import secure_filename
                import pandas as pd
                
                data_type = request.form.get('data_type')
                file = request.files.get('file')
                
                if not data_type or not file:
                    flash('Please select a data type and upload a file', 'error')
                    return redirect(url_for('admin_import_data', auth_token=request.form.get('auth_token')))
                
                # Secure the filename and save file
                filename = secure_filename(file.filename)
                upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                # Ensure upload directory exists
                os.makedirs(upload_folder, exist_ok=True)
                filepath = os.path.join(upload_folder, filename)
                file.save(filepath)
                
                # Read file to get columns for mapping
                try:
                    file_ext = '.' + filename.rsplit('.', 1)[-1].lower()
                    if file_ext == '.csv':
                        df = pd.read_csv(filepath, nrows=5)  # Read first 5 rows for preview
                    else:
                        df = pd.read_excel(filepath, nrows=5)
                    
                    columns = df.columns.tolist()
                except Exception as e:
                    os.remove(filepath)  # Clean up file on error
                    flash(f'Error reading file: {str(e)}', 'error')
                    return redirect(url_for('admin_import_data', auth_token=request.form.get('auth_token')))
                
                # Get current user
                user = current_user
                
                # Create import job
                job = ImportJob(
                    user_id=user.id,
                    data_type=data_type,
                    filename=filename,
                    status='Mapping'
                )
                db.session.add(job)
                db.session.commit()
                
                # Redirect to mapping page
                return redirect(url_for('admin_import_map', job_id=job.id, auth_token=request.form.get('auth_token')))
                
            except Exception as e:
                flash(f'Upload failed: {str(e)}', 'error')
                return redirect(url_for('admin_import_data', auth_token=request.form.get('auth_token')))
        
        # Get current user from context  
        user = current_user
        
        # Get recent import jobs for current user
        import_jobs = ImportJob.query.filter_by(user_id=user.id).order_by(ImportJob.created_at.desc()).limit(10).all() if user else []
        
        # Show upload form by default
        return render_template('admin/import_upload.html', 
                             import_jobs=import_jobs,
                             auth_token=request.args.get('auth_token'))

    @app.route('/admin/import-data/map/<int:job_id>', methods=['GET', 'POST'])
    @admin_required
    def admin_import_map(job_id):
        """Column mapping interface"""
        from models import ImportJob
        import pandas as pd
        import io
        
        job = ImportJob.query.get_or_404(job_id)
        
        if request.method == 'POST':
            # Save mapping and redirect to execute
            mapping = {}
            field_definitions = {
                'Problem': ['title', 'description', 'priority', 'reporter_email', 'department_name', 'status', 'impact', 'urgency'],
                'BusinessCase': ['title', 'summary', 'case_type', 'cost_estimate', 'benefit_estimate', 'submitter_email', 'department_name', 'status'],
                'Project': ['name', 'description', 'project_manager_email', 'department_name', 'status', 'budget', 'start_date', 'end_date']
            }
            
            for field in field_definitions[job.data_type]:
                column = request.form.get(field)
                if column:
                    mapping[column] = field
            
            job.mapping = mapping
            job.status = 'Importing'
            db.session.commit()
            
            return redirect(url_for('admin_import_execute', job_id=job.id, auth_token=request.form.get('auth_token')))
        
        # Read file to get actual columns
        import pandas as pd
        
        try:
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            filepath = os.path.join(upload_folder, job.filename)
            file_ext = '.' + job.filename.rsplit('.', 1)[-1].lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(filepath, nrows=5)
            else:
                df = pd.read_excel(filepath, nrows=5)
            
            columns = df.columns.tolist()
        except Exception as e:
            columns = ['Error reading file columns: ' + str(e)]
        
        return render_template('admin/import_map.html', 
                             job=job, 
                             columns=columns,
                             auth_token=request.args.get('auth_token'))

    @app.route('/admin/import-data/result/<int:job_id>')
    @admin_required
    def admin_import_result(job_id):
        """Show import results"""
        from models import ImportJob
        
        job = ImportJob.query.get_or_404(job_id)
        
        return render_template('admin/import_result.html', 
                             job=job,
                             auth_token=request.args.get('auth_token'))

    @app.route('/admin/sample-data')
    @admin_required
    def admin_sample_data():
        """Sample data download page"""
        return render_template('admin/sample_data.html', 
                             auth_token=request.args.get('auth_token'))

    @app.route('/admin/import-data/status')
    @admin_required
    def admin_import_status_overview():
        """Get import job status for real-time updates"""
        from models import ImportJob
        from flask import jsonify
        
        jobs = ImportJob.query.order_by(ImportJob.created_at.desc()).limit(10).all()
        
        job_data = []
        for job in jobs:
            job_data.append({
                'id': job.id,
                'data_type': job.data_type,
                'filename': job.filename,
                'status': job.status,
                'rows_success': job.rows_success or 0,
                'rows_failed': job.rows_failed or 0,
                'created_at': format_datetime(job.created_at) if job.created_at else None
            })
        
        return jsonify(job_data)

    @app.route('/admin/import-data/upload', methods=['POST'])
    @admin_required
    def admin_import_upload():
        """Handle file upload and preview"""
        try:
            if 'file' not in request.files:
                return jsonify({'success': False, 'error': 'No file uploaded'}), 400
            
            file = request.files['file']
            data_type = request.form.get('data_type')
            
            if not data_type or data_type not in ['Problem', 'BusinessCase', 'Project']:
                return jsonify({'success': False, 'error': 'Invalid data type'}), 400
            
            if not file or not file.filename:
                return jsonify({'success': False, 'error': 'No file selected'}), 400
            
            # Validate file type
            allowed_extensions = {'.csv', '.xlsx', '.xls'}
            file_ext = '.' + file.filename.rsplit('.', 1)[-1].lower()
            
            if file_ext not in allowed_extensions:
                return jsonify({'success': False, 'error': f'Unsupported file format. Allowed: {", ".join(allowed_extensions)}'}), 400
            
            # Read file preview using pandas
            import pandas as pd
            import io
            
            file_content = file.read()
            file.seek(0)  # Reset file pointer
            
            if file_ext == '.csv':
                df = pd.read_csv(io.BytesIO(file_content), nrows=10)
            else:
                df = pd.read_excel(io.BytesIO(file_content), nrows=10)
            
            # Create import job
            from models import ImportJob
            job = ImportJob(
                user_id=current_user.id,
                data_type=data_type,
                filename=file.filename,
                status='Mapping'
            )
            db.session.add(job)
            db.session.commit()
            
            # Define field requirements
            field_requirements = {
                'Problem': {
                    'required': ['title', 'description'],
                    'optional': ['priority', 'reporter_email', 'department_name', 'status', 'impact', 'urgency']
                },
                'BusinessCase': {
                    'required': ['title', 'summary'],
                    'optional': ['case_type', 'cost_estimate', 'benefit_estimate', 'submitter_email', 'department_name', 'status']
                },
                'Project': {
                    'required': ['name', 'description'],
                    'optional': ['project_manager_email', 'department_name', 'status', 'budget', 'start_date', 'target_end_date']
                }
            }
            
            return jsonify({
                'success': True,
                'job_id': job.id,
                'preview': {
                    'columns': df.columns.tolist(),
                    'preview_data': df.to_dict('records'),
                    'total_rows': len(df)
                },
                'field_requirements': field_requirements[data_type]
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/admin/import-data/mapping', methods=['POST'])
    @admin_required
    def admin_import_mapping():
        """Save column mapping configuration"""
        try:
            job_id = request.form.get('job_id')
            mapping_data = request.form.get('mapping')
            
            if not job_id or not mapping_data:
                return jsonify({'success': False, 'error': 'Missing job ID or mapping data'}), 400
            
            # Parse mapping JSON
            import json
            mapping = json.loads(mapping_data)
            
            # Save mapping
            from models import ImportJob
            job = ImportJob.query.get(int(job_id))
            if not job:
                return jsonify({'success': False, 'error': 'Job not found'}), 404
            
            job.mapping = mapping
            job.status = 'Importing'
            db.session.commit()
            
            return jsonify({'success': True, 'message': 'Mapping saved successfully'})
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/admin/import-data/execute/<int:job_id>', methods=['GET', 'POST'])
    @admin_required  
    def admin_import_execute(job_id):
        """Execute the import process"""
        from flask import jsonify
        
        try:
            # For GET requests, show a simple execution page
            if request.method == 'GET':
                from models import ImportJob
                job = ImportJob.query.get_or_404(job_id)
                return render_template('admin/import_execute.html', 
                                     job=job,
                                     auth_token=request.args.get('auth_token'))
            
            # For POST requests, execute the import
            from models import ImportJob, Problem, BusinessCase, Project, User, Department
            import pandas as pd
            
            job = ImportJob.query.get(int(job_id))
            if not job or not job.mapping:
                return jsonify({'success': False, 'error': 'Invalid job or missing mapping'}), 400
            
            # Update status to Importing if not already
            if job.status != 'Importing':
                job.status = 'Importing'
                db.session.commit()
            
            # Read full file from uploaded location
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            filepath = os.path.join(upload_folder, job.filename)
            
            # Check if file exists
            if not os.path.exists(filepath):
                # Try to find file in static/sample_data for sample files
                if job.filename.startswith('sample_data_'):
                    filepath = os.path.join('static/sample_data', job.filename)
                if not os.path.exists(filepath):
                    return jsonify({'success': False, 'error': f'File not found: {job.filename}'}), 400
            
            file_ext = '.' + job.filename.rsplit('.', 1)[-1].lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(filepath)
            else:
                df = pd.read_excel(filepath)
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Get row inclusion preferences from the form (from checkboxes)
            import_rows = set()
            for key in request.form.keys():
                if key.startswith('include_row_'):
                    row_num = int(key.replace('include_row_', ''))
                    import_rows.add(row_num - 1)  # Convert to 0-based index
            
            print(f"Import rows selected: {import_rows}")
            print(f"Processing {len(df)} rows with mapping: {job.mapping}")
            
            # Process each row
            for index, row in df.iterrows():
                # Check if this row was selected for import
                if import_rows and index not in import_rows:
                    print(f"Skipping row {index + 1} - excluded by user (duplicate or unchecked)")
                    errors.append({
                        'row': index + 1,
                        'error': 'Skipped duplicate',
                        'type': 'duplicate_skipped'
                    })
                    error_count += 1
                    continue
                try:
                    # Prepare data based on mapping
                    data = {}
                    
                    for csv_field, model_field in job.mapping.items():
                        if csv_field in row and pd.notna(row[csv_field]):
                            value = str(row[csv_field]).strip()
                            
                            # Handle special field mappings using helper functions
                            if 'email' in csv_field:
                                user_id = _lookup_user_id(value)
                                if user_id:
                                    data[model_field.replace('_email', '_id')] = user_id
                            elif 'department_name' in csv_field:
                                dept_id = _lookup_department_id(value)
                                if dept_id:
                                    data['department_id'] = dept_id
                            elif model_field == 'problem_id':
                                problem_id = _lookup_problem_id(value)
                                if problem_id:
                                    data[model_field] = problem_id
                            elif model_field == 'case_id':
                                case_id = _lookup_case_id(value)
                                if case_id:
                                    data[model_field] = case_id
                            else:
                                # Direct field mapping
                                data[model_field] = value
                    
                    print(f"Row {index + 2} data: {data}")
                    
                    # Add required fields that might be missing
                    current_user = current_user
                    
                    if job.data_type == 'Problem':
                        # Add default values for required fields if not provided
                        if 'reported_by' not in data:
                            data['reported_by'] = current_user.id
                        if 'department_id' not in data:
                            # Use current user's department or first available department
                            if current_user.department_id:
                                data['department_id'] = current_user.department_id
                            else:
                                first_dept = Department.query.filter_by(organization_id=current_user.organization_id).first()
                                if first_dept:
                                    data['department_id'] = first_dept.id
                        
                        # Handle enum fields properly with proper casting
                        if 'priority' in data:
                            from models import PriorityEnum
                            priority_value = str(data['priority']).strip()
                            try:
                                # Try direct enum value mapping
                                if priority_value in ['Critical', 'Urgent']:
                                    data['priority'] = PriorityEnum.High  # Map Critical to High
                                else:
                                    data['priority'] = PriorityEnum(priority_value)
                            except ValueError:
                                # Fallback to Medium if invalid value
                                data['priority'] = PriorityEnum.Medium
                        
                        if 'status' in data:
                            from models import StatusEnum
                            status_value = str(data['status']).strip()
                            try:
                                # Map common variations to enum values
                                status_mapping = {
                                    'In Progress': StatusEnum.InProgress,
                                    'In_Progress': StatusEnum.InProgress,
                                    'InProgress': StatusEnum.InProgress,
                                    'On Hold': StatusEnum.OnHold,
                                    'On_Hold': StatusEnum.OnHold,
                                    'OnHold': StatusEnum.OnHold,
                                    'Resolved': StatusEnum.Resolved,
                                    'Open': StatusEnum.Open
                                }
                                data['status'] = status_mapping.get(status_value, StatusEnum(status_value))
                            except ValueError:
                                data['status'] = StatusEnum.Open
                        
                        if 'impact' in data:
                            from models import ImpactEnum
                            impact_value = str(data['impact']).strip()
                            try:
                                data['impact'] = ImpactEnum(impact_value)
                            except ValueError:
                                data['impact'] = ImpactEnum.Medium
                        
                        if 'urgency' in data:
                            from models import UrgencyEnum
                            urgency_value = str(data['urgency']).strip()
                            try:
                                data['urgency'] = UrgencyEnum(urgency_value)
                            except ValueError:
                                data['urgency'] = UrgencyEnum.Medium
                    
                    elif job.data_type == 'BusinessCase':
                        # Add default values for business case required fields
                        if 'submitted_by' not in data:
                            data['submitted_by'] = current_user.id
                        if 'department_id' not in data:
                            if current_user.department_id:
                                data['department_id'] = current_user.department_id
                            else:
                                first_dept = Department.query.filter_by(organization_id=current_user.organization_id).first()
                                if first_dept:
                                    data['department_id'] = first_dept.id
                        
                        # Handle enum fields for business case
                        if 'case_type' in data:
                            from models import CaseTypeEnum
                            case_type_value = data['case_type']
                            if case_type_value == 'Reactive':
                                data['case_type'] = CaseTypeEnum.Reactive
                            else:
                                data['case_type'] = CaseTypeEnum.Proactive
                        
                        if 'status' in data:
                            from models import StatusEnum
                            status_value = data['status']
                            if status_value == 'In Progress':
                                data['status'] = StatusEnum.InProgress
                            elif status_value == 'Approved':
                                data['status'] = StatusEnum.Approved
                            elif status_value == 'Rejected':
                                data['status'] = StatusEnum.Rejected
                            else:
                                data['status'] = StatusEnum.Open
                        
                        # Convert numeric fields
                        if 'cost_estimate' in data:
                            try:
                                data['cost_estimate'] = float(data['cost_estimate'])
                            except (ValueError, TypeError):
                                data['cost_estimate'] = 0.0
                        
                        if 'benefit_estimate' in data:
                            try:
                                data['benefit_estimate'] = float(data['benefit_estimate'])
                            except (ValueError, TypeError):
                                data['benefit_estimate'] = 0.0
                    
                    # Create record based on data type
                    if job.data_type == 'Problem':
                        from models import Problem, StatusEnum
                        
                        # Generate code if not present
                        if 'code' not in data or not data['code']:
                            last_problem = Problem.query.order_by(Problem.id.desc()).first()
                            next_id = (last_problem.id + 1) if last_problem else 1
                            data['code'] = f"P{next_id:04d}"
                        
                        # Set required fields with defaults
                        data['reported_by'] = data.get('reported_by') or current_user.id
                        data['created_by'] = data.get('created_by') or current_user.id
                        
                        # Handle department assignment - use fallback if not provided
                        # Department can be updated later via problem editing
                        if not data.get('department_id'):
                            # Try current user's department first
                            if current_user.department_id:
                                data['department_id'] = current_user.department_id
                            else:
                                # Use first available department as fallback
                                first_dept = Department.query.filter_by(organization_id=current_user.organization_id).first()
                                if first_dept:
                                    data['department_id'] = first_dept.id
                                else:
                                    # Skip this row if no departments exist at all
                                    error_details.append({
                                        'row': idx + 1,
                                        'error': 'No departments available in system - please create departments first'
                                    })
                                    continue
                        
                        # Remove any old dept_id field to avoid conflicts (Problem model uses department_id)
                        if 'org_unit_id' in data:
                            del data['org_unit_id']
                        
                        # Handle status mapping for Problems using StatusEnum (since ProblemStatusEnum doesn't exist)
                        if 'status' in data:
                            status_value = str(data['status']).strip()
                            status_mapping = {
                                'open': StatusEnum.Open,
                                'investigating': StatusEnum.InProgress,
                                'resolved': StatusEnum.Resolved,
                                'closed': StatusEnum.Resolved,
                                'in progress': StatusEnum.InProgress
                            }
                            mapped_status = status_mapping.get(status_value.lower())
                            if mapped_status:
                                data['status'] = mapped_status
                            else:
                                data['status'] = StatusEnum.Open
                        else:
                            data['status'] = StatusEnum.Open
                        
                        print(f"Creating Problem with data: {data}")
                        record = Problem(**data)
                    elif job.data_type == 'BusinessCase':
                        from models import BusinessCase
                        
                        # Generate code if not present
                        if 'code' not in data or not data['code']:
                            last_case = BusinessCase.query.order_by(BusinessCase.id.desc()).first()
                            next_id = (last_case.id + 1) if last_case else 1
                            data['code'] = f"C{next_id:04d}"
                        
                        # Set required fields with defaults
                        data['created_by'] = data.get('created_by') or current_user.id
                        
                        # Remove submitted_by field as BusinessCase model doesn't have it
                        if 'submitted_by' in data:
                            del data['submitted_by']
                        
                        # Handle department assignment (BusinessCase uses department_id)
                        if not data.get('department_id'):
                            if current_user.department_id:
                                data['department_id'] = current_user.department_id
                            else:
                                first_dept = Department.query.filter_by(organization_id=current_user.organization_id).first()
                                if first_dept:
                                    data['department_id'] = first_dept.id
                        
                        # Remove department_id field if present (BusinessCase uses dept_id)
                        if 'department_id' in data:
                            del data['department_id']
                        
                        # Map 'summary' field to 'description' if present (BusinessCase uses description, not summary)
                        if 'summary' in data:
                            data['description'] = data.pop('summary')
                        
                        print(f"Creating BusinessCase with data: {data}")
                        record = BusinessCase(**data)
                    elif job.data_type == 'Project':
                        from models import Project, StatusEnum
                        
                        # Generate code if not present
                        if 'code' not in data or not data['code']:
                            last_project = Project.query.order_by(Project.id.desc()).first()
                            next_id = (last_project.id + 1) if last_project else 1
                            data['code'] = f"PRJ{next_id:04d}"
                        
                        # Set required fields with defaults
                        data['created_by'] = data.get('created_by') or current_user.id
                        data['project_manager_id'] = data.get('project_manager_id') or current_user.id
                        
                        # Handle department assignment - use fallback if not provided
                        # Department can be updated later via project editing
                        if not data.get('department_id'):
                            # Try current user's department first
                            if current_user.department_id:
                                data['department_id'] = current_user.department_id
                            else:
                                # Use first available department as fallback
                                first_dept = Department.query.filter_by(organization_id=current_user.organization_id).first()
                                if first_dept:
                                    data['department_id'] = first_dept.id
                                # If no departments exist, leave as None (nullable field)
                        
                        # Remove any old dept_id field to avoid conflicts
                        if 'org_unit_id' in data:
                            del data['org_unit_id']
                        data['department_id'] = data.get('department_id') or current_user.department_id
                        
                        # Handle status mapping for Projects using StatusEnum
                        if 'status' in data:
                            status_value = str(data['status']).strip()
                            # Map common project status values to StatusEnum
                            status_mapping = {
                                'planning': StatusEnum.Open,
                                'active': StatusEnum.InProgress, 
                                'in progress': StatusEnum.InProgress,
                                'on hold': StatusEnum.OnHold,  # Use OnHold instead of Draft
                                'completed': StatusEnum.Resolved,
                                'cancelled': StatusEnum.Resolved
                            }
                            mapped_status = status_mapping.get(status_value.lower())
                            if mapped_status:
                                data['status'] = mapped_status
                            else:
                                data['status'] = StatusEnum.Open  # Default to Open
                        else:
                            data['status'] = StatusEnum.Open
                        
                        print(f"Creating Project with data: {data}")
                        record = Project(**data)
                    else:
                        raise ValueError(f"Unknown data type: {job.data_type}")
                    
                    db.session.add(record)
                    db.session.flush()  # Flush to get any database errors before commit
                    success_count += 1
                    print(f"Successfully added record {index + 2}")
                
                except Exception as e:
                    db.session.rollback()
                    error_count += 1
                    print(f"Error importing row {index + 2}: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    errors.append({
                        'row': index + 2,  # +2 because pandas is 0-indexed and we have header
                        'error': str(e)
                    })
            
            # Commit all successful records
            if success_count > 0:
                db.session.commit()
            
            # Update job status
            job.rows_success = success_count
            job.rows_failed = error_count
            job.error_details = errors
            job.status = 'Complete' if error_count == 0 else 'Failed'
            db.session.commit()
            
            print(f"Import completed: {success_count} success, {error_count} errors")
            
            # Clean up uploaded file after processing
            try:
                upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                filepath = os.path.join(upload_folder, job.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
            except Exception:
                pass  # Don't fail import if cleanup fails
            
            message = f"Import completed: {success_count} records imported successfully"
            if error_count > 0:
                message += f", {error_count} failed"
            
            return jsonify({
                'success': True,
                'message': message,
                'redirect_url': url_for('admin_import_result', job_id=job.id, auth_token=request.form.get('auth_token'))
            })
        
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500

    @app.route('/admin/import-data/status/<int:job_id>')
    @admin_required
    def admin_import_job_status(job_id):
        """Get import job status"""
        from models import ImportJob
        
        job = ImportJob.query.get(job_id)
        if not job:
            return jsonify({'success': False, 'error': 'Job not found'}), 404
        
        return jsonify({
            'success': True,
            'status': job.status,
            'rows_success': job.rows_success,
            'rows_failed': job.rows_failed,
            'error_details': job.error_details
        })
    
    @app.route('/admin/workflows/library/<int:library_id>/import', methods=['POST'])
    @login_required
    @admin_required
    def import_from_library(library_id):
        """Import workflow from library and return edit modal data"""
        library_workflow = WorkflowLibrary.query.filter_by(id=library_id, organization_id=current_user.organization_id).first_or_404()
        
        # Create new workflow template from library
        new_workflow = WorkflowTemplate(
            name=f"My - {library_workflow.name}",
            description=library_workflow.description,
            definition=library_workflow.definition,
            is_active=False,  # Start as inactive for customization
            created_by=current_user.id
        )
        
        db.session.add(new_workflow)
        db.session.commit()
        log_action(f"Imported workflow from library: {library_workflow.name}")
        
        # Return JSON response with new workflow ID for immediate editing
        from flask import jsonify
        return jsonify({
            'success': True,
            'workflow_id': new_workflow.id,
            'name': new_workflow.name,
            'description': new_workflow.description,
            'definition': new_workflow.definition,
            'is_active': new_workflow.is_active
        })
    
    @app.route('/admin/workflows/import', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def import_workflow():
        """Import workflow from library with enhanced functionality"""
        from models import WorkflowLibrary, WorkflowTemplate
        
        libraries = WorkflowLibrary.query.order_by(WorkflowLibrary.category, WorkflowLibrary.name).all()
        
        if request.method == 'POST':
            lib_id = request.form.get('library_id', type=int)
            custom_name = request.form.get('custom_name', '').strip()
            
            if not lib_id:
                return render_template('admin/import_workflow.html', 
                                     libraries=libraries, 
                                     error="Please select a workflow to import.")
            
            lib = WorkflowLibrary.query.filter_by(id=lib_id, organization_id=current_user.organization_id).first_or_404()
            
            # Use custom name if provided, otherwise default naming
            workflow_name = custom_name if custom_name else f"My - {lib.name}"
            
            # Check for duplicate names
            existing = WorkflowTemplate.query.filter_by(name=workflow_name, organization_id=current_user.organization_id).first()
            if existing:
                return render_template('admin/import_workflow.html', 
                                     libraries=libraries, 
                                     error=f"A workflow named '{workflow_name}' already exists.")
            
            wt = WorkflowTemplate(
                name=workflow_name,
                description=lib.description,
                definition=lib.definition,
                is_active=False,  # Start inactive for customization
                created_by=current_user.id
            )
            
            db.session.add(wt)
            db.session.commit()
            log_action(f"Imported workflow from library: {lib.name} as {workflow_name}")
            
            from flask import flash
            flash(f"Workflow '{workflow_name}' imported successfully! You can now customize it.", "success")
            return redirect(url_for('admin_workflows'))
        
        return render_template('admin/import_workflow.html', libraries=libraries)
    
    @app.route('/admin/org-structure/import', methods=['GET','POST'])
    @login_required
    @admin_required
    def import_org_chart():
        """Simple org chart import with GET/POST pattern"""
        if request.method == 'POST':
            try:
                from models import Department, User
                import pandas as pd
                
                file = request.files['file']
                if not file or file.filename == '':
                    from flask import flash
                    flash('No file selected', 'error')
                    return render_template('admin/import_org_chart.html')
                
                # Save file temporarily to read with pandas
                import tempfile
                import os
                with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as temp_file:
                    file.save(temp_file.name)
                    df = pd.read_csv(temp_file.name)  # expects columns: name,parent_name,manager_email
                
                success_count = 0
                for _, row in df.iterrows():
                    try:
                        unit = Department.query.filter_by(name=row['name']).first() or Department(name=row['name'])
                        
                        if pd.notna(row.get('parent_name')):
                            parent = Department.query.filter_by(name=row['parent_name']).first()
                            if parent:
                                unit.parent = parent
                        
                        if pd.notna(row.get('manager_email')):
                            mgr = User.query.filter_by(email=row['manager_email']).first()
                            if mgr:
                                unit.manager = mgr
                        
                        db.session.add(unit)
                        success_count += 1
                        
                    except Exception as e:
                        print(f"Error processing row: {str(e)}")
                        continue
                
                # Clean up temp file
                if 'temp_file' in locals():
                    os.unlink(temp_file.name)
                
                db.session.commit()
                log_action('org_chart_imported_simple', f'Imported {success_count} org units')
                
                from flask import flash
                flash(f'Org chart imported successfully. {success_count} units processed.', 'success')
                return redirect('/admin/org-structure')
                
            except Exception as e:
                db.session.rollback()
                log_action('org_chart_import_error', f'Error: {str(e)}')
                from flask import flash
                flash(f'Import error: {str(e)}', 'error')
                return render_template('admin/import_org_chart.html')
        
        return render_template('admin/import_org_chart.html')
    
    @app.route('/admin/org-structure/<int:id>/edit', methods=['POST'])
    @login_required
    @admin_required
    def edit_org_unit(id):
        """Edit a department"""
        try:
            from models import Department, User
            
            unit = Department.query.get_or_404(id)
            
            # Get form data
            name = request.form.get('name', '').strip()
            manager_id = request.form.get('manager_id')
            parent_id = request.form.get('parent_id')
            
            if not name:
                from flask import flash
                flash('Unit name is required', 'error')
                return redirect('/admin/org-structure')
            
            # Update unit
            unit.name = name
            
            # Set manager
            if manager_id and manager_id != '':
                manager = User.query.get(int(manager_id))
                if manager:
                    unit.manager = manager
                else:
                    unit.manager = None
            else:
                unit.manager = None
            
            # Set parent (with circular reference check)
            if parent_id and parent_id != '':
                parent = Department.query.get(int(parent_id))
                if parent and parent.id != unit.id:
                    # Check for circular reference
                    current = parent
                    while current:
                        if current.id == unit.id:
                            from flask import flash
                            flash('Cannot set parent - would create circular reference', 'error')
                            return redirect('/admin/org-structure')
                        current = current.parent
                    unit.parent = parent
                else:
                    unit.parent = None
            else:
                unit.parent = None
            
            db.session.commit()
            log_action('org_unit_updated', f'Updated org unit: {unit.name}')
            
            from flask import flash
            flash(f'Organizational unit "{unit.name}" updated successfully', 'success')
            return redirect('/admin/org-structure')
            
        except Exception as e:
            db.session.rollback()
            log_action('org_unit_update_error', f'Error updating org unit {id}: {str(e)}')
            from flask import flash
            flash(f'Error updating organizational unit: {str(e)}', 'error')
            return redirect('/admin/org-structure')
    
    @app.route('/admin/org-structure', methods=['GET'])
    @login_required
    @admin_required
    def view_org_chart():
        """View departmental chart with hierarchical display"""
        try:
            from models import Department, User
            from flask import request, jsonify
            
            # Get all departments with their hierarchy - filter by organization
            from flask_login import current_user
            org_id = current_user.organization_id if current_user.is_authenticated else None
            
            if org_id:
                departments = Department.query.filter_by(organization_id=org_id).all()
                roots = Department.query.filter_by(parent_id=None, organization_id=org_id).all()
            else:
                departments = []
                roots = []
            
            print(f"DEBUG: Total departments: {len(departments)}")
            print(f"DEBUG: Root departments: {len(roots)}")
            for root in roots:
                print(f"DEBUG: Root department: {root.name} (ID: {root.id})")
            
            # Serialize departments for JavaScript
            def serialize_department(dept):
                try:
                    # Get manager info
                    manager_info = None
                    if dept.manager_id:
                        manager = User.query.get(dept.manager_id)
                        if manager:
                            manager_info = {
                                'id': manager.id,
                                'name': manager.name or manager.email,
                                'email': manager.email
                            }
                    
                    # Get children recursively
                    children_data = []
                    children = Department.query.filter_by(parent_id=dept.id).all()
                    for child in children:
                        children_data.append(serialize_department(child))
                    
                    result = {
                        'id': dept.id,
                        'name': dept.name,
                        'parent_id': dept.parent_id,
                        'manager': manager_info,
                        'children': children_data
                    }
                    print(f"DEBUG: Serialized department {dept.name}: {result}")
                    return result
                except Exception as e:
                    log_action('serialize_error', f'Error serializing department {dept.id}: {str(e)}')
                    print(f"DEBUG: Error serializing {dept.name}: {str(e)}")
                    return {
                        'id': dept.id,
                        'name': dept.name,
                        'parent_id': dept.parent_id,
                        'manager': None,
                        'children': []
                    }
            
            try:
                serialized_roots = [serialize_department(root) for root in roots]
                log_action('org_structure_serialized', f'Serialized {len(serialized_roots)} root departments from {len(roots)} total')
                print(f"DEBUG: Found {len(roots)} root departments, serialized {len(serialized_roots)} departments")
                print(f"DEBUG: Root departments: {[r.name for r in roots]}")
                print(f"DEBUG: Serialized data: {serialized_roots}")
            except Exception as e:
                log_action('serialization_error', f'Error serializing roots: {str(e)}')
                print(f"DEBUG: Serialization error: {str(e)}")
                serialized_roots = []
            
            serialized_departments = [{
                'id': dept.id,
                'name': dept.name,
                'parent_id': dept.parent_id,
                'manager_id': dept.manager_id
            } for dept in departments]
            
            # Check if this is a JSON request
            if request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'departments': serialized_departments,
                    'roots': serialized_roots
                })
            
            return render_template('admin/org_structure.html', 
                                departments=serialized_departments,
                                roots=serialized_roots,
                                current_user=current_user)
        except Exception as e:
            log_action('org_structure_view_error', f'Error: {str(e)}')
            if request.headers.get('Accept') == 'application/json':
                return jsonify({'error': str(e)}), 500
            return f"Error loading org structure page: {str(e)}", 500
    
    @app.route('/api/users', methods=['GET'])
    @login_required
    def api_users():
        """API endpoint to get users for dropdowns - ORGANIZATION FILTERED"""
        try:
            from models import User
            # CRITICAL SECURITY FIX: Filter users by organization_id
            users = User.query.filter_by(organization_id=current_user.organization_id).all()
            users_data = []
            for user in users:
                # Use actual User model structure: name, email, role
                users_data.append({
                    'id': user.id,
                    'email': user.email,
                    'name': getattr(user, 'name', ''),
                    'role': str(getattr(user, 'role', 'Unknown'))
                })
            return jsonify(users_data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/sample_org_chart.csv', methods=['GET'])
    def download_sample_org_chart():
        """Download sample organizational chart CSV file"""
        try:
            from flask import send_file
            import os
            
            # Get the file path
            file_path = os.path.join(os.getcwd(), 'sample_org_chart.csv')
            
            if os.path.exists(file_path):
                return send_file(
                    file_path,
                    as_attachment=True,
                    download_name='sample_org_chart.csv',
                    mimetype='text/csv'
                )
            else:
                return "Sample file not found", 404
        except Exception as e:
            return f"Error downloading file: {str(e)}", 500

    @app.route('/admin/org-structure/<int:unit_id>/delete', methods=['POST'])
    @login_required
    def delete_org_unit(unit_id):
        """Delete organizational unit with cascade deletion of children"""
        try:
            from models import Department, db
            
            # Find the unit to delete
            unit = Department.query.get_or_404(unit_id)
            unit_name = unit.name
            
            # Delete the unit (cascade will handle children)
            db.session.delete(unit)
            db.session.commit()
            
            # Log the action
            log_action(f"Deleted department: {unit_name} (ID: {unit_id})")
            
            return "Unit deleted successfully", 200
            
        except Exception as e:
            db.session.rollback()
            return f"Error deleting unit: {str(e)}", 500

    @app.route('/admin/org-structure/create', methods=['POST'])
    @login_required 
    def create_org_unit():
        """Create new organizational unit"""
        try:
            from models import Department, User, db
            from flask import request
            
            name = request.form.get('name', '').strip()
            manager_id = request.form.get('manager_id')
            parent_id = request.form.get('parent_id')
            
            if not name:
                return "Department name is required", 400
            
            # Create new department with organization_id
            from flask_login import current_user
            org_id = current_user.organization_id if current_user.is_authenticated else None
            
            new_unit = Department(name=name, organization_id=org_id)
            
            # Set manager if provided - ensure manager is from same organization
            if manager_id and manager_id.strip():
                manager = User.query.filter_by(id=manager_id, organization_id=org_id).first()
                if manager:
                    new_unit.manager_id = manager.id
            
            # Set parent if provided - ensure parent is from same organization
            if parent_id and parent_id.strip():
                parent = Department.query.filter_by(id=parent_id, organization_id=org_id).first()
                if parent:
                    new_unit.parent_id = parent.id
            
            db.session.add(new_unit)
            db.session.commit()
            
            # Log the action
            log_action(f"Created department: {name}")
            
            return redirect(url_for('view_org_chart'))
            
        except Exception as e:
            db.session.rollback()
            return f"Error creating department: {str(e)}", 500

    @app.route('/admin/org-structure/export')
    @login_required
    def export_org_structure():
        """Export organizational structure to CSV"""
        try:
            from models import Department
            from flask import Response
            import io
            import csv
            
            print(f"🔧 Export Debug: User {current_user.email} requesting org structure export")
            
            # Get all departments for current organization
            units = Department.query.filter_by(organization_id=current_user.organization_id).all()
            print(f"🔧 Export Debug: Found {len(units)} units for organization {current_user.organization_id}")
            
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow(['name', 'manager_email', 'parent_name'])
            
            # Write units data
            for unit in units:
                manager_email = unit.manager.email if unit.manager else ''
                parent_name = unit.parent.name if unit.parent else ''
                writer.writerow([unit.name, manager_email, parent_name])
            
            # Create response
            csv_data = output.getvalue()
            output.close()
            
            print(f"🔧 Export Debug: Generated CSV with {len(csv_data)} characters")
            
            return Response(
                csv_data,
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=organizational_structure.csv'}
            )
            
        except Exception as e:
            print(f"🔧 Export Error: {str(e)}")
            return f"Error exporting organizational structure: {str(e)}", 500
    
    @app.route('/admin/export-test')
    def export_test():
        """Simple export test route"""
        try:
            print(f"🔧 Export Test Debug: current_user.is_authenticated = {current_user.is_authenticated}")
            if current_user.is_authenticated:
                return f"✓ Export test successful for user: {current_user.email} in org: {current_user.organization_id}"
            else:
                return f"❌ Export test failed: User not authenticated. Session: {dict(session)}"
        except Exception as e:
            return f"❌ Export test failed: {str(e)}"
    
    @app.route('/admin/simple-test')
    def simple_test():
        """Simplest possible admin test route"""
        return "<h1>Simple Test Route Working</h1><p>This route exists and is accessible.</p>"
    
    @app.route('/admin/test-export-no-auth')
    def test_export_no_auth():
        """Test route with no authentication"""
        return f"""
        <h1>Export Test - No Auth Required</h1>
        <p>This route works without authentication.</p>
        <p>Current time: {datetime.now()}</p>
        <p>If you can see this, basic routing is working.</p>
        <p><a href="/admin/">Back to Admin</a></p>
        """
    
    @app.route('/admin/debug-data-export', methods=['GET'])
    def debug_data_export():
        """Debug version with manual auth check"""
        try:
            print(f"🚀 DEBUG EXPORT: Route accessed!")
            print(f"🚀 DEBUG EXPORT: User authenticated: {current_user.is_authenticated}")
            
            if not current_user.is_authenticated:
                return redirect(url_for('auth.login', next=request.url))
            
            print(f"🚀 DEBUG EXPORT: User: {current_user.email}")
            print(f"🚀 DEBUG EXPORT: Role: {current_user.role}")
            
            # Manual admin check
            if current_user.role.name != 'Admin':
                return "Access denied - Admin role required", 403
            
            from models import Department, Problem, BusinessCase, Project
            
            org_id = current_user.organization_id
            departments_count = Department.query.filter_by(organization_id=org_id).count()
            problems_count = Problem.query.filter_by(organization_id=org_id).count()
            business_cases_count = BusinessCase.query.filter_by(organization_id=org_id).count()
            projects_count = Project.query.filter_by(organization_id=org_id).count()
            
            return f"""
            <h1>✅ Data Export - Working!</h1>
            <h2>Organization Statistics for {current_user.email}</h2>
            <p><strong>Organization ID:</strong> {org_id}</p>
            <ul>
                <li><strong>Departments:</strong> {departments_count}</li>
                <li><strong>Problems:</strong> {problems_count}</li>
                <li><strong>Business Cases:</strong> {business_cases_count}</li>
                <li><strong>Projects:</strong> {projects_count}</li>
            </ul>
            <p><a href="/admin/" class="btn btn-primary">Back to Admin Dashboard</a></p>
            <p>✅ Authentication and Data Export Working Correctly!</p>
            """
        except Exception as e:
            print(f"🚀 DEBUG EXPORT ERROR: {str(e)}")
            return f"<h1>Debug Data Export Error</h1><p>{str(e)}</p>"
    
    # Register the main data export route with explicit methods
    @app.route('/admin/data-export', methods=['GET', 'POST'])
    @login_required  
    @admin_required
    def admin_data_export():
        """Admin data export page with multiple export options"""
        try:
            print(f"🚀 DATA EXPORT ROUTE ACCESSED SUCCESSFULLY! User: {current_user.email}")
            print(f"🚀 DATA EXPORT: User authenticated: {current_user.is_authenticated}")
            print(f"🚀 DATA EXPORT: User role: {current_user.role}")
            print(f"🚀 DATA EXPORT: Organization ID: {current_user.organization_id}")
            
            from models import Department, Problem, BusinessCase, Project
            
            # Get counts for current organization
            org_id = current_user.organization_id
            print(f"🔧 Data Export Debug: Organization ID: {org_id}")
            
            export_stats = {
                'departments': Department.query.filter_by(organization_id=org_id).count(),
                'problems': Problem.query.filter_by(organization_id=org_id).count(),
                'business_cases': BusinessCase.query.filter_by(organization_id=org_id).count(),
                'projects': Project.query.filter_by(organization_id=org_id).count()
            }
            
            print(f"🔧 Data Export Debug: Export stats: {export_stats}")
            
            return render_template('admin/data_export.html', 
                                 export_stats=export_stats,
                                 user=current_user)
        except Exception as e:
            print(f"🔧 Data Export Error: {str(e)}")
            return f"<h1>Data Export Error</h1><p>{str(e)}</p>"

    print(f"🔧 Route Registration Debug: Data export route registered successfully")

    @app.route('/admin/org-reports', methods=['GET'])
    @login_required
    def org_reports():
        """Organizational reports dashboard"""
        try:
            from models import Department
            
            # Get all departments
            all_units = Department.query.filter_by(organization_id=current_user.organization_id).all()
            
            # Calculate statistics
            total_units = len(all_units)
            managed_units = len([u for u in all_units if u.manager_id])
            root_units = len([u for u in all_units if u.parent_id is None])
            
            # Calculate max depth
            max_depth = 0
            for unit in all_units:
                depth = unit.get_hierarchy_level()
                if depth > max_depth:
                    max_depth = depth
            
            org_stats = {
                'total_units': total_units,
                'managed_units': managed_units,
                'root_units': root_units,
                'max_depth': max_depth + 1  # Add 1 for display (level 0 = level 1)
            }
            
            # Prepare hierarchy data with full path
            hierarchy_data = []
            for unit in all_units:
                unit_data = unit
                unit_data.level = unit.get_hierarchy_level()
                unit_data.full_path_str = unit.get_full_path()
                hierarchy_data.append(unit_data)
            
            # Sort by hierarchy level and name
            hierarchy_data.sort(key=lambda x: (x.level, x.name))
            
            # Get tree structure for chart
            org_tree_data = serialize_org_unit_tree()
            
            return render_template('admin/org_reports.html',
                                org_stats=org_stats,
                                hierarchy_data=hierarchy_data,
                                org_tree_data=org_tree_data[0] if org_tree_data else {})
                                
        except Exception as e:
            return f"Error loading organizational reports: {str(e)}", 500

    @app.route('/admin/org-reports/download/<format>', methods=['GET'])
    @login_required
    def download_org_report(format):
        """Download organizational report in specified format"""
        try:
            from models import Department
            from flask import Response, send_file, render_template
            import io
            import csv
            
            all_units = Department.query.filter_by(organization_id=current_user.organization_id).all()
            
            if format == 'csv':
                # Generate CSV report
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Write header
                writer.writerow(['Unit Name', 'Hierarchy Level', 'Manager Name', 'Manager Email', 
                               'Parent Unit', 'Direct Reports', 'Full Path'])
                
                # Write data
                for unit in all_units:
                    writer.writerow([
                        unit.name,
                        unit.get_hierarchy_level() + 1,
                        unit.manager.name if unit.manager else '',
                        unit.manager.email if unit.manager else '',
                        unit.parent.name if unit.parent else 'Root Unit',
                        len(unit.children),
                        unit.get_full_path()
                    ])
                
                csv_data = output.getvalue()
                output.close()
                
                return Response(
                    csv_data,
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=organizational_report.csv'}
                )
                
            elif format == 'excel':
                # Generate Excel report using openpyxl
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill
                    import tempfile
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Organizational Structure"
                    
                    # Headers
                    headers = ['Unit Name', 'Hierarchy Level', 'Manager Name', 'Manager Email', 
                              'Parent Unit', 'Direct Reports', 'Full Path']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    # Data
                    for row, unit in enumerate(all_units, 2):
                        ws.cell(row=row, column=1, value=unit.name)
                        ws.cell(row=row, column=2, value=unit.get_hierarchy_level() + 1)
                        ws.cell(row=row, column=3, value=unit.manager.name if unit.manager else '')
                        ws.cell(row=row, column=4, value=unit.manager.email if unit.manager else '')
                        ws.cell(row=row, column=5, value=unit.parent.name if unit.parent else 'Root Unit')
                        ws.cell(row=row, column=6, value=len(unit.children))
                        ws.cell(row=row, column=7, value=unit.get_full_path())
                    
                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Save to temporary file
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        wb.save(tmp.name)
                        tmp.seek(0)
                        
                        return send_file(
                            tmp.name,
                            as_attachment=True,
                            download_name='organizational_report.xlsx',
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                        
                except ImportError:
                    return "Excel export requires openpyxl package", 500
                    
            elif format == 'pdf':
                # Generate PDF report using WeasyPrint
                try:
                    from weasyprint import HTML, CSS
                    import tempfile
                    
                    # Create HTML content
                    html_content = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="UTF-8">
                        <title>Organizational Structure Report</title>
                        <style>
                            body {{ font-family: Arial, sans-serif; margin: 20px; }}
                            h1 {{ color: #333; border-bottom: 2px solid #333; padding-bottom: 10px; }}
                            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
                            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                            th {{ background-color: #f2f2f2; font-weight: bold; }}
                            .stats {{ display: flex; justify-content: space-around; margin: 20px 0; }}
                            .stat-box {{ text-align: center; padding: 10px; border: 1px solid #ddd; }}
                            .level-indent {{ margin-left: 20px; }}
                        </style>
                    </head>
                    <body>
                        <h1>Organizational Structure Report</h1>
                        
                        <div class="stats">
                            <div class="stat-box">
                                <h3>{len(all_units)}</h3>
                                <p>Total Units</p>
                            </div>
                            <div class="stat-box">
                                <h3>{max([unit.get_hierarchy_level() for unit in all_units]) + 1 if all_units else 0}</h3>
                                <p>Max Levels</p>
                            </div>
                            <div class="stat-box">
                                <h3>{len([u for u in all_units if u.manager_id])}</h3>
                                <p>Managed Units</p>
                            </div>
                            <div class="stat-box">
                                <h3>{len([u for u in all_units if u.parent_id is None])}</h3>
                                <p>Root Units</p>
                            </div>
                        </div>
                        
                        <table>
                            <tr>
                                <th>Unit Name</th>
                                <th>Level</th>
                                <th>Manager</th>
                                <th>Parent Unit</th>
                                <th>Reports</th>
                                <th>Full Path</th>
                            </tr>
                    """
                    
                    # Sort units by hierarchy
                    sorted_units = sorted(all_units, key=lambda x: (x.get_hierarchy_level(), x.name))
                    
                    for unit in sorted_units:
                        level_indent = "margin-left: {}px;".format(unit.get_hierarchy_level() * 20)
                        html_content += f"""
                            <tr>
                                <td style="{level_indent}">{'└─ ' if unit.get_hierarchy_level() > 0 else ''}{unit.name}</td>
                                <td>{unit.get_hierarchy_level() + 1}</td>
                                <td>{unit.manager.name if unit.manager else 'No Manager'}</td>
                                <td>{unit.parent.name if unit.parent else 'Root Unit'}</td>
                                <td>{len(unit.children)}</td>
                                <td>{unit.get_full_path()}</td>
                            </tr>
                        """
                    
                    html_content += """
                        </table>
                    </body>
                    </html>
                    """
                    
                    # Generate PDF
                    pdf_bytes = HTML(string=html_content).write_pdf()
                    
                    return Response(
                        pdf_bytes,
                        mimetype='application/pdf',
                        headers={'Content-Disposition': 'attachment; filename=organizational_report.pdf'}
                    )
                    
                except ImportError:
                    return "PDF export requires weasyprint package", 500
                    
            else:
                return "Invalid format", 400
                
        except Exception as e:
            return f"Error generating report: {str(e)}", 500

    @app.route('/admin/org-reports/chart/<format>', methods=['POST'])
    @login_required
    def download_org_chart(format):
        """Download organizational chart in specified format"""
        try:
            from flask import request, Response
            import json
            
            # Get SVG data from request
            data = request.get_json()
            svg_data = data.get('svg_data', '')
            
            if format == 'svg':
                return Response(
                    svg_data,
                    mimetype='image/svg+xml',
                    headers={'Content-Disposition': 'attachment; filename=organizational_chart.svg'}
                )
                
            elif format == 'png':
                try:
                    # Convert SVG to PNG using cairosvg
                    import cairosvg
                    png_data = cairosvg.svg2png(bytestring=svg_data.encode('utf-8'))
                    
                    return Response(
                        png_data,
                        mimetype='image/png',
                        headers={'Content-Disposition': 'attachment; filename=organizational_chart.png'}
                    )
                except ImportError:
                    # Fallback: return SVG if cairosvg not available
                    return Response(
                        svg_data,
                        mimetype='image/svg+xml',
                        headers={'Content-Disposition': 'attachment; filename=organizational_chart.svg'}
                    )
                    
            else:
                return "Invalid chart format", 400
                
        except Exception as e:
            return f"Error generating chart: {str(e)}", 500

    def serialize_org_unit_tree():
        """Serialize departments into tree structure for charts"""
        from models import Department
        
        # Get all root departments (departments without parents)
        root_units = Department.query.filter_by(parent_id=None).all()
        
        def serialize_unit(unit):
            # Get children for this department
            children = Department.query.filter_by(parent_id=unit.id).all()
            
            return {
                'id': unit.id,
                'name': unit.name,
                'manager': {
                    'id': unit.manager.id,
                    'name': unit.manager.name,
                    'email': unit.manager.email
                } if unit.manager else None,
                'children': [serialize_unit(child) for child in children]
            }
        
        return [serialize_unit(unit) for unit in root_units]

    # Help Center Management Routes
    @app.route('/admin/help-center')
    @login_required
    @admin_required
    def admin_help_center():
        """Help Center management dashboard"""
        categories = HelpCategory.query.filter_by(organization_id=current_user.organization_id).order_by(HelpCategory.sort_order, HelpCategory.name).all()
        articles = HelpArticle.query.filter_by(organization_id=current_user.organization_id).join(HelpCategory).order_by(HelpCategory.sort_order, HelpArticle.sort_order, HelpArticle.title).all()
        
        # Add article count to categories
        for category in categories:
            category.article_count = len([a for a in articles if a.category_id == category.id])
        
        return render_template('admin/help_center.html', categories=categories, articles=articles)
    
    @app.route('/admin/help-center/categories/create', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_create_help_category():
        """Create new help category"""
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            sort_order = request.form.get('sort_order', 0)
            
            if not name:
                flash('Category name is required', 'error')
                return redirect(url_for('admin_create_help_category'))
            
            # Check for duplicate name
            existing = HelpCategory.query.filter_by(name=name, organization_id=current_user.organization_id).first()
            if existing:
                flash('Category name already exists', 'error')
                return redirect(url_for('admin_create_help_category'))
            
            try:
                category = HelpCategory(
                    organization_id=current_user.organization_id,
                    name=name,
                    sort_order=int(sort_order) if sort_order else 0
                )
                db.session.add(category)
                db.session.commit()
                
                log_action(f"Created help category: {name}")
                flash('Help category created successfully', 'success')
                return redirect(url_for('admin_help_center'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'Error creating category: {str(e)}', 'error')
                return redirect(url_for('admin_create_help_category'))
        
        return render_template('admin/help_category_form.html', category=None)
    
    @app.route('/admin/help-center/categories/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_edit_help_category(id):
        """Edit help category"""
        category = HelpCategory.query.filter_by(id=id, organization_id=current_user.organization_id).first_or_404()
        
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            sort_order = request.form.get('sort_order', 0)
            
            if not name:
                flash('Category name is required', 'error')
                return redirect(url_for('admin_edit_help_category', id=id))
            
            # Check for duplicate name (excluding current category)
            existing = HelpCategory.query.filter(HelpCategory.name == name, HelpCategory.id != id).first()
            if existing:
                flash('Category name already exists', 'error')
                return redirect(url_for('admin_edit_help_category', id=id))
            
            try:
                category.name = name
                category.sort_order = int(sort_order) if sort_order else 0
                db.session.commit()
                
                log_action(f"Updated help category: {name}")
                flash('Help category updated successfully', 'success')
                return redirect(url_for('admin_help_center'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating category: {str(e)}', 'error')
        
        return render_template('admin/help_category_form.html', category=category)
    
    @app.route('/admin/help-center/categories/<int:id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def admin_delete_help_category(id):
        """Delete help category"""
        category = HelpCategory.query.filter_by(id=id, organization_id=current_user.organization_id).first_or_404()
        
        # Check if category has articles
        if category.articles:
            flash(f'Cannot delete category "{category.name}" - it contains {len(category.articles)} articles', 'error')
            return redirect(url_for('admin_help_center'))
        
        try:
            category_name = category.name
            db.session.delete(category)
            db.session.commit()
            
            log_action(f"Deleted help category: {category_name}")
            flash('Help category deleted successfully', 'success')
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting category: {str(e)}', 'error')
        
        return redirect(url_for('admin_help_center'))
    
    @app.route('/admin/help-center/articles/create', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_create_help_article():
        """Create new help article"""
        categories = HelpCategory.query.filter_by(organization_id=current_user.organization_id).order_by(HelpCategory.sort_order, HelpCategory.name).all()
        
        if request.method == 'POST':
            category_id = request.form.get('category_id')
            title = request.form.get('title', '').strip()
            content = request.form.get('content', '').strip()
            sort_order = request.form.get('sort_order', 0)
            
            if not title:
                flash('Article title is required', 'error')
                return redirect(url_for('admin_create_help_article'))
            
            if not content:
                flash('Article content is required', 'error')
                return redirect(url_for('admin_create_help_article'))
            
            if not category_id:
                flash('Category is required', 'error')
                return redirect(url_for('admin_create_help_article'))
            
            try:
                article = HelpArticle(
                    organization_id=current_user.organization_id,
                    category_id=int(category_id),
                    title=title,
                    content=content,
                    sort_order=int(sort_order) if sort_order else 0,
                    created_by=current_user.id
                )
                
                # Generate slug
                article.slug = article.generate_slug()
                
                db.session.add(article)
                db.session.commit()
                
                log_action(f"Created help article: {title}")
                flash('Help article created successfully', 'success')
                return redirect(url_for('admin_help_center'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'Error creating article: {str(e)}', 'error')
        
        return render_template('admin/help_article_form.html', article=None, categories=categories)
    
    @app.route('/admin/help-center/articles/<int:id>/edit', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_edit_help_article(id):
        """Edit help article"""
        article = HelpArticle.query.filter_by(id=id, organization_id=current_user.organization_id).first_or_404()
        categories = HelpCategory.query.filter_by(organization_id=current_user.organization_id).order_by(HelpCategory.sort_order, HelpCategory.name).all()
        
        if request.method == 'POST':
            category_id = request.form.get('category_id')
            title = request.form.get('title', '').strip()
            content = request.form.get('content', '').strip()
            sort_order = request.form.get('sort_order', 0)
            
            if not title:
                flash('Article title is required', 'error')
                return redirect(url_for('admin_edit_help_article', id=id))
            
            if not content:
                flash('Article content is required', 'error')
                return redirect(url_for('admin_edit_help_article', id=id))
            
            if not category_id:
                flash('Category is required', 'error')
                return redirect(url_for('admin_edit_help_article', id=id))
            
            try:
                old_title = article.title
                article.category_id = int(category_id)
                article.title = title
                article.content = content
                article.sort_order = int(sort_order) if sort_order else 0
                
                # Regenerate slug if title changed
                if title != old_title:
                    article.slug = article.generate_slug()
                
                db.session.commit()
                
                log_action(f"Updated help article: {title}")
                flash('Help article updated successfully', 'success')
                return redirect(url_for('admin_help_center'))
            
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating article: {str(e)}', 'error')
        
        return render_template('admin/help_article_form.html', article=article, categories=categories)
    
    @app.route('/admin/preferences-demo')
    @login_required 
    @admin_required
    def admin_preferences_demo():
        """Demonstration page for organization preferences system"""
        return render_template('admin/preferences_demo.html')
    
    @app.route('/admin/organization-settings', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def admin_organization_settings():
        """Comprehensive organization preferences settings page"""
        # Get or create organization settings
        org_settings = OrganizationSettings.get_organization_settings()
        
        if request.method == 'POST':
            # Update organization settings from form data
            org_settings.timezone = request.form.get('timezone', 'UTC')
            org_settings.currency = request.form.get('currency', 'USD')
            org_settings.date_format = request.form.get('date_format', 'ISO')
            org_settings.time_format = request.form.get('time_format', '%H:%M:%S')
            org_settings.default_theme = request.form.get('default_theme', 'light')
            org_settings.updated_at = datetime.utcnow()
            org_settings.updated_by = current_user.id
            
            try:
                db.session.commit()
                log_action('UPDATE_ORG_SETTINGS', f'Updated organization preferences: currency={org_settings.currency}, theme={org_settings.default_theme}')
                flash('Organization preferences updated successfully!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating settings: {str(e)}', 'error')
            
            return redirect(url_for('admin_organization_settings'))
        
        # Prepare data for template
        timezone_choices = get_timezone_choices()
        currency_choices = [
            ('USD', 'US Dollar ($)'),
            ('EUR', 'Euro (€)'),
            ('GBP', 'British Pound (£)'),
            ('CAD', 'Canadian Dollar (C$)'),
            ('AUD', 'Australian Dollar (A$)'),
            ('JPY', 'Japanese Yen (¥)'),
            ('CNY', 'Chinese Yuan (¥)'),
            ('INR', 'Indian Rupee (₹)')
        ]
        
        date_format_choices = [
            ('ISO', '2024-12-31 (ISO format)'),
            ('%m/%d/%Y', '12/31/2024 (US format)'),
            ('%d/%m/%Y', '31/12/2024 (EU format)'),
            ('%d-%m-%Y', '31-12-2024 (EU dashes)'),
            ('%B %d, %Y', 'December 31, 2024 (Full month)')
        ]
        
        time_format_choices = [
            ('%H:%M:%S', '23:59:59 (24-hour)'),
            ('%H:%M', '23:59 (24-hour, no seconds)'),
            ('%I:%M:%S %p', '11:59:59 PM (12-hour)'),
            ('%I:%M %p', '11:59 PM (12-hour, no seconds)')
        ]
        
        theme_choices = [
            ('light', 'Light Theme'),
            ('dark', 'Dark Theme')
        ]
        
        return render_template('admin/organization_settings.html',
                             org_settings=org_settings,
                             timezone_choices=timezone_choices,
                             currency_choices=currency_choices,
                             date_format_choices=date_format_choices,
                             time_format_choices=time_format_choices,
                             theme_choices=theme_choices)
    
    @app.route('/admin/help-center/articles/<int:id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def admin_delete_help_article(id):
        """Delete help article"""
        article = HelpArticle.query.filter_by(id=id, organization_id=current_user.organization_id).first_or_404()
        
        try:
            article_title = article.title
            db.session.delete(article)
            db.session.commit()
            
            log_action(f"Deleted help article: {article_title}")
            flash('Help article deleted successfully', 'success')
        
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting article: {str(e)}', 'error')
        
        return redirect(url_for('admin_help_center'))

    @app.route('/admin/run-triage')
    @login_required
    @admin_required
    def run_triage_now():
        """Execute triage rules manually"""
        try:
            from services.triage_engine import apply_rules
            applied_count = apply_rules()
            
            log_action(f"Manually executed triage rules - {applied_count} actions applied")
            flash(f'Triage rules executed successfully. {applied_count} actions applied.', 'success')
            
        except Exception as e:
            flash(f'Error executing triage rules: {str(e)}', 'error')
            
        return redirect(url_for('admin_dashboard'))

    @app.route('/admin/triage-rules')
    @login_required
    @admin_required
    def admin_triage_rules():
        """Triage rules management interface"""
        from models import TriageRule
        
        rules = TriageRule.query.order_by(TriageRule.created_at.desc()).all()
        
        # Get available targets, fields, operators, and actions for dropdowns
        targets = ['Epic', 'BusinessCase', 'Project']
        operators = ['=', '>', '<', '>=', '<=', 'contains', 'days_ago']
        actions = ['auto_approve', 'flag', 'notify_admin', 'escalate']
        
        # Field mappings for each target
        field_mappings = {
            'Epic': ['title', 'description', 'status', 'created_at', 'estimated_effort'],
            'BusinessCase': ['title', 'description', 'status', 'created_at', 'estimated_cost', 'classification'],
            'Project': ['title', 'description', 'status', 'created_at', 'budget', 'priority']
        }
        
        return render_template('admin/triage_rules.html',
                             rules=rules,
                             targets=targets,
                             operators=operators,
                             actions=actions,
                             field_mappings=field_mappings)

    @app.route('/admin/triage-rules/create', methods=['POST'])
    @login_required
    @admin_required
    def admin_create_triage_rule():
        """Create new triage rule"""
        from models import TriageRule
        
        try:
            name = request.form.get('name', '').strip()
            target = request.form.get('target', '').strip()
            field = request.form.get('field', '').strip()
            operator = request.form.get('operator', '').strip()
            value = request.form.get('value', '').strip()
            action = request.form.get('action', '').strip()
            message = request.form.get('message', '').strip()
            
            if not all([name, target, field, operator, value, action]):
                flash('All fields except message are required', 'error')
                return redirect(url_for('admin_triage_rules'))
            
            rule = TriageRule(
                name=name,
                target=target,
                field=field,
                operator=operator,
                value=value,
                action=action,
                message=message or None,
                created_by=current_user.id
            )
            
            db.session.add(rule)
            db.session.commit()
            
            log_action(f"Created triage rule: {name}")
            flash(f'Triage rule "{name}" created successfully', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating triage rule: {str(e)}', 'error')
        
        return redirect(url_for('admin_triage_rules'))

    @app.route('/admin/triage-rules/<int:rule_id>/toggle', methods=['POST'])
    @login_required
    @admin_required
    def admin_toggle_triage_rule(rule_id):
        """Toggle triage rule active status"""
        from models import TriageRule
        
        try:
            rule = TriageRule.query.filter_by(id=rule_id, organization_id=current_user.organization_id).first_or_404()
            rule.active = not rule.active
            db.session.commit()
            
            status = "activated" if rule.active else "deactivated"
            log_action(f"Triage rule '{rule.name}' {status}")
            flash(f'Triage rule "{rule.name}" {status}', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating triage rule: {str(e)}', 'error')
        
        return redirect(url_for('admin_triage_rules'))

    @app.route('/admin/triage-rules/<int:rule_id>/delete', methods=['POST'])
    @login_required
    @admin_required
    def admin_delete_triage_rule(rule_id):
        """Delete triage rule"""
        from models import TriageRule
        
        try:
            rule = TriageRule.query.filter_by(id=rule_id, organization_id=current_user.organization_id).first_or_404()
            rule_name = rule.name
            db.session.delete(rule)
            db.session.commit()
            
            log_action(f"Deleted triage rule: {rule_name}")
            flash(f'Triage rule "{rule_name}" deleted successfully', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error deleting triage rule: {str(e)}', 'error')
        
        return redirect(url_for('admin_triage_rules'))

    @app.route('/admin/triage-rules/<int:rule_id>/test', methods=['POST'])
    @login_required
    @admin_required
    def admin_test_triage_rule(rule_id):
        """Test triage rule (dry run)"""
        from models import TriageRule
        from services.triage_engine import test_rule
        
        try:
            rule = TriageRule.query.filter_by(id=rule_id, organization_id=current_user.organization_id).first_or_404()
            count, results = test_rule(rule)
            
            if isinstance(results, str):
                flash(f'Rule test failed: {results}', 'error')
            else:
                log_action(f"Tested triage rule '{rule.name}' - would affect {count} entities")
                flash(f'Rule "{rule.name}" would affect {count} entities', 'info')
                
                # Store test results in session for display
                session['triage_test_results'] = {
                    'rule_name': rule.name,
                    'count': count,
                    'samples': results[:5]  # Limit to 5 samples
                }
            
        except Exception as e:
            flash(f'Error testing triage rule: {str(e)}', 'error')
        
        return redirect(url_for('admin_triage_rules'))

    # Removed conflicting simple notification interface - using detailed interface at /admin/notifications/ instead



    print("✓ Admin routes initialized successfully")

def get_timezone_choices():
    """Get timezone choices for the organization settings form"""
    return [
        ('UTC', 'UTC (Coordinated Universal Time)'),
        ('US/Eastern', 'US Eastern Time'),
        ('US/Central', 'US Central Time'), 
        ('US/Mountain', 'US Mountain Time'),
        ('US/Pacific', 'US Pacific Time'),
        ('Europe/London', 'London (GMT/BST)'),
        ('Europe/Paris', 'Paris (CET/CEST)'),
        ('Europe/Berlin', 'Berlin (CET/CEST)'),
        ('Asia/Tokyo', 'Tokyo (JST)'),
        ('Asia/Shanghai', 'Shanghai (CST)'),
        ('Asia/Kolkata', 'India (IST)'),
        ('Australia/Sydney', 'Sydney (AEST/AEDT)')
    ]